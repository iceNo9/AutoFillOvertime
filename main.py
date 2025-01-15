import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from datetime import datetime
from tqdm import tqdm

def dynamic_string_prompt(base_string):
    # 使用 tqdm 来控制循环次数，设定循环次数为无限次，模拟重复循环
    print(base_string, end="", flush=True)

    # 旋转符号列表（可以根据需求修改）
    symbols = ['|', '/','-', '\\']

    # 无限循环，动态更新字符串
    while True:
        for symbol in symbols:
            sys.stdout.write("\r" + base_string + symbol)  # \r 会覆盖前面的输出
            sys.stdout.flush()  # 刷新输出
            time.sleep(0.5)  # 每半秒更新一次

        # 等待回车键输入
        user_input = input()  # 等待用户按回车键，按回车后退出

        # 如果按下回车键则退出循环
        if user_input == '':
            break

def read_overtime_data(file_path):
    # 读取 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    
    # 获取名为 "detail" 的工作表
    if 'detail' in wb.sheetnames:
        sheet = wb['detail']
    else:
        print("没有找到名为 'detail' 的工作表！")
        return None

    # 字典，用于存储不同类型的加班数据
    overtime_data = {
        "普通加班": [],
        "公休加班": [],
        "节日加班": []
    }

    # 获取表头，假设表头在第一行
    headers = [cell.value for cell in sheet[1]]
    
    # 查找加班时长列和其他需要的列
    overtime_duration_idx = headers.index("加班时长")
    overtime_start_idx = headers.index("加班开始时间")
    overtime_end_idx = headers.index("加班结束时间")
    overtime_reason_idx = headers.index("加班原因")
    overtime_type_idx = headers.index("类型")
    
    # 遍历所有数据行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        overtime_duration = row[overtime_duration_idx]
        overtime_start = row[overtime_start_idx]
        overtime_end = row[overtime_end_idx]
        overtime_reason = row[overtime_reason_idx]
        overtime_type = row[overtime_type_idx]

        # 筛选加班时长>=1的数据
        if overtime_duration and overtime_duration >= 1:
            # 解析加班开始和结束时间
            start_time = datetime.strptime(str(overtime_start), "%Y-%m-%d %H:%M:%S")
            end_time = datetime.strptime(str(overtime_end), "%Y-%m-%d %H:%M:%S")
            
            # 提取日期和时间
            start_date = start_time.date()
            start_time_only = start_time.time()
            end_date = end_time.date()
            end_time_only = end_time.time()

            # 创建一个字典存储每一条加班记录
            overtime_record = {
                "加班开始日期": start_date,
                "加班开始时间": start_time_only,
                "加班结束日期": end_date,
                "加班结束时间": end_time_only,
                "加班原因": overtime_reason,
                "加班类型": overtime_type
            }

            # 将加班类型映射为分类
            if overtime_type == "工作日":
                overtime_data["普通加班"].append(overtime_record)
            elif overtime_type == "公休日":
                overtime_data["公休加班"].append(overtime_record)
            elif overtime_type == "节假日":
                overtime_data["节日加班"].append(overtime_record)

    return overtime_data

def main(file_path):
    try:
        # 获取当前目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"当前目录: {current_dir}")

        # 拼接 Chrome 浏览器路径
        chrome_path = os.path.join(current_dir, 'chrome-win64', 'chrome.exe')
        print(f"Chrome路径: {chrome_path}")

        driver_path = os.path.join(current_dir, 'chromedriver-win64', 'chromedriver.exe')
        service = Service(driver_path)
        
        # 设置 Chrome 浏览器选项
        chrome_options = Options()
        chrome_options.binary_location = chrome_path
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--allow-insecure-localhost")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-setuid-sandbox")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # 禁用浏览器自动化标志
        # chrome_options.add_argument("--headless")  # 如果你需要无头模式，可以取消注释这一行

        # 启动 Chrome 浏览器
        driver = webdriver.Chrome(service=service, options=chrome_options)

        # 添加反自动化检测处理
        def add_antidetect():
            """注入代码来绕过反自动化检测"""
            driver.execute_script('''
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
            ''')
            driver.execute_script('navigator.__defineGetter__("userAgent", function(){return "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36";});')

        # 调用反自动化处理
        # add_antidetect()

        # 打开网页
        driver.get('http://10.10.0.20')  # 替换为你需要访问的 URL

        dynamic_string_prompt("打开HR41流程界面后，按Enter继续")

        # 获取全部窗口句柄
        window_handles= driver.window_handles

        # 标志，是否找到目标标签
        found = False

        # 遍历所有标签页，找到标题包含 "HR41" 的标签
        for handle in window_handles:
            driver.switch_to.window(handle)  # 切换到当前标签页
            if "HR41" in driver.title:  # 检查标签的标题
                print(f"切换到标签：{driver.title}")
                found = True
                break  # 找到后就退出循环

        if not found:
            print("没有找到带有 'HR41' 的标签页 即将退出！")
            driver.quit()  # 关闭浏览器
            sys.exit("程序退出：未找到目标标签页！")  # 退出程序

        # 读取文件，显示加班数据
        overtime_data = read_overtime_data(file_path)

        while True:
            # 提示信息
            print("****菜单选项****：")
            print("1. 普通加班")
            print("2. 公休加班")
            print("3. 节日加班")
            print("4. 退出")

            # 遍历加班数据字典，检查每种加班类型的数量
            available_types = []
            if len(overtime_data['普通加班']) > 0:
                available_types.append('1')
                print(f"普通加班：{len(overtime_data['普通加班'])} 条")
            if len(overtime_data['公休加班']) > 0:
                available_types.append('2')
                print(f"公休加班：{len(overtime_data['公休加班'])} 条")
            if len(overtime_data['节日加班']) > 0:
                available_types.append('3')
                print(f"节日加班：{len(overtime_data['节日加班'])} 条")

            # 用户输入
            user_input = input("请输入数字：")

            # 处理用户输入
            if user_input == '1' and '1' in available_types:
                list_data = overtime_data['普通加班']
                pass
            elif user_input == '2' and '2' in available_types:
                list_data = overtime_data['公休加班']
                pass
            elif user_input == '3' and '3' in available_types:
                list_data = overtime_data['节日加班']
                pass
            elif user_input == '4' and '4' in available_types:
                break
            else:
                print("输入错误或没有可用的加班类型，请检查并重新输入。")
                continue
            
            element_add = driver.find_element(By.CSS_SELECTOR, "#addbutton0")  # 替换为实际按钮的选择器
            # input("新增按钮定位完成，按Enter继续...")

            row_index = 4
            item_index = 35
            detail_index = 8

            for data in list_data:
                date_elements = []
                time_elements = []
                error_count = 0
                
                driver.execute_script("arguments[0].click();", element_add)
                time.sleep(0.1)
                # input("新增条目结束，按Enter继续...")

                parent_element = driver.find_element(By.CSS_SELECTOR, f"#oTable0 > tbody > tr:nth-child({row_index})")
                print(parent_element.get_attribute("class"))
                print(parent_element.get_attribute("data-rowindex"))

                while True:
                    if (2 == len(date_elements)) and (2 == len(time_elements)):
                        break
                    else:
                        error_count = error_count + 1 
                        # 日期
                        date_elements = parent_element.find_elements(By.CSS_SELECTOR, ".picker-icon.icon-coms-New-schedule.cursor-pointer")
                        # 时间
                        time_elements = parent_element.find_elements(By.CSS_SELECTOR, ".picker-icon.icon-coms-overtime.cursor-pointer")
                        if 200 < error_count:
                            raise ValueError("可识别元素不完整")
                

                start_date_element = date_elements[0]
                end_date_element = date_elements[1]
                start_time_element = time_elements[0]
                end_time_element = time_elements[1]

                # 日期
                driver.execute_script("arguments[0].click();", start_date_element)

                # 输入日期后回车
                element = driver.find_element(By.CSS_SELECTOR, f"body > div:nth-child({item_index}) > div > div > div > div:nth-child(1) > div:nth-child(2) > input") 
                item_index = item_index + 1
                driver.execute_script(f"arguments[0].value = '{data['加班开始日期']}';", element)
                element.send_keys(Keys.ENTER)

                # 起始时间
                driver.execute_script("arguments[0].click();", start_time_element)

                # 输入起始时间后回车
                element = driver.find_element(By.CSS_SELECTOR, f"body > div:nth-child({item_index}) > div > div > div > div.ant-time-picker-panel-input-wrap > input") 
                item_index = item_index + 1
                driver.execute_script(f"arguments[0].value = '{data['加班开始时间']}';", element)
                element.send_keys(Keys.ENTER)

                # 结束时间
                driver.execute_script("arguments[0].click();", end_time_element)

                # 输入结束时间后回车
                element = driver.find_element(By.CSS_SELECTOR, f"body > div:nth-child({item_index}) > div > div > div > div.ant-time-picker-panel-input-wrap > input") 
                item_index = item_index + 1
                driver.execute_script(f"arguments[0].value = '{data['加班结束时间']}';", element)
                element.send_keys(Keys.ENTER)

                # 输入详细后回车
                element = driver.find_element(By.CSS_SELECTOR, f"#oTable0 > tbody > tr:nth-child({row_index}) > td:nth-child({detail_index}) > div > div > input") 
                driver.execute_script(f"arguments[0].value = '{data['加班原因']}';", element)
                driver.execute_script("arguments[0].click();", element)

                row_index = row_index + 1

            dynamic_string_prompt("输入完成，按Enter回到菜单...")

        # 关闭浏览器
        driver.quit()

    except Exception as e:
        print(f"异常：{e}")

if __name__ == "__main__":
    VER = "V1.0-"
    DATE = "20250115"    
    commit_message = "feea:初版作成，支持HR41加班单自动填充；"

    print(f"版本: {VER}{DATE}")
    print("版本修改说明: " + commit_message)

    # 从命令行获取文件路径
    if len(sys.argv) > 1:
        file_path = sys.argv[1]  # 获取传入的文件路径
        main(file_path)
    else:
        print("请提供文件路径参数。")